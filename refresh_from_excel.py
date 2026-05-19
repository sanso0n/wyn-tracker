#!/usr/bin/env python3
"""
WYN AD Tracker — refresco desde Excel
=====================================
PROTOCOLO:
  - La fuente PRIMARIA durante la semana es el `Datos_parciales_WYN_vs_RIVAL.xlsx`
    de la carpeta más reciente dentro del ciclo en curso (ej. Gold Group Mayo).
    Cada día agrego los puntos Lun-(hoy) por jugador y los muestro como
    ranking provisional.
  - El `Clasificacion_WYN.xlsx` (master) aporta el agregado oficial al cierre
    de cada semana: hojas `Week1 WYN` (o WeekN), `Datos Cofres`, `Cofres — Ranking`,
    `Historial`. Si esa hoja existe y está poblada, sus datos prevalecen sobre el
    cálculo vivo para esa semana concreta.
  - Las competiciones cerradas (Diamond Group Abril) se leen del Excel archivado
    en su carpeta correspondiente.

Uso:
    python3 refresh_from_excel.py            # refresco completo
    python3 refresh_from_excel.py --preview  # solo enseña el resumen

No escribe nunca fuera de App WYN/.
"""
from __future__ import annotations
import sys, json, statistics, argparse, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl. Instálalo con:  pip3 install --user openpyxl")
    sys.exit(1)

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent

# ───────────────────────────────────────────────────────────
# PROTECCIÓN DE VERSIONES CONGELADAS
# ───────────────────────────────────────────────────────────
# Si este script se ejecuta accidentalmente desde dentro de una carpeta
# congelada (v1_CONGELADA_NO_TOCAR, v2_CONGELADA_..., etc.), abortamos
# inmediatamente sin escribir nada. La regla: cualquier ruta que contenga
# un componente con '_CONGELADA_' o que sea exactamente 'vN' / 'vN_*' es
# de solo lectura.
def _is_frozen_path(p: Path) -> bool:
    for part in p.parts:
        if '_CONGELADA_' in part: return True
        if len(part) >= 2 and part[0] == 'v' and part[1].isdigit():
            # 'v1', 'v1_anything', 'v23_anything'
            return True
    return False

if _is_frozen_path(HERE):
    print(f'⛔ Esta carpeta está congelada y no debe modificarse: {HERE}')
    print('   Si quieres regenerar la app, ejecuta refresh_from_excel.py desde la carpeta App WYN/, no desde una v#.')
    sys.exit(2)
MASTER = ROOT / "Clasificacion_WYN.xlsx"
ABRIL  = ROOT / "Diamond Group Abril" / "Clasificacion_WYN Abril.xlsx"
OUT_JSON  = HERE / "current-week-extended.json"
OUT_HIST  = HERE / "history.json"
OUT_HTML  = HERE / "index.html"
TEMPLATE  = HERE / "template.html"

# Carpeta del ciclo en curso. Si cambia (ej. Gold Group Junio), editar aquí
# o pasarlo como argumento --cycle "Gold Group Mayo".
DEFAULT_CYCLE_DIR_NAME = "Gold Group Mayo"

# Reference time: Apocalypse Time = UTC.
# 3 operational zones based on the player's local UTC offset.
COUNTRY_ZONE = {
    # ─────── EMEA · Apocalypse −1 to +4 (Europe, Africa, Middle East) ───────
    'Great Britain':'EMEA','Ireland':'EMEA','Portugal':'EMEA','Iceland':'EMEA',
    'Spain':'EMEA','France':'EMEA','Germany':'EMEA','Italy':'EMEA',
    'Netherlands':'EMEA','Belgium':'EMEA','Luxembourg':'EMEA','Monaco':'EMEA',
    'Switzerland':'EMEA','Austria':'EMEA','Denmark':'EMEA','Norway':'EMEA','Sweden':'EMEA',
    'Poland':'EMEA','Czech Republic':'EMEA','Slovakia':'EMEA','Hungary':'EMEA',
    'Slovenia':'EMEA','Croatia':'EMEA','Serbia':'EMEA','Bosnia':'EMEA','Albania':'EMEA',
    'Greece':'EMEA','Romania':'EMEA','Bulgaria':'EMEA','Ukraine':'EMEA',
    'Estonia':'EMEA','Latvia':'EMEA','Lithuania':'EMEA','Finland':'EMEA',
    'Türkiye':'EMEA','Turkey':'EMEA','Israel':'EMEA','Georgia':'EMEA',
    'Russia':'EMEA',
    'Algeria':'EMEA','Morocco':'EMEA','Tunisia':'EMEA','Egypt':'EMEA',
    'Nigeria':'EMEA','Kenya':'EMEA','South Africa':'EMEA','Ethiopia':'EMEA',
    'Saudi Arabia':'EMEA','UAE':'EMEA','Qatar':'EMEA',
    # ─────── APAC · Apocalypse +5 to +12 (Asia-Pacific) ───────
    'Pakistan':'APAC','India':'APAC','Indien':'APAC','Bangladesh':'APAC',
    'Thailand':'APAC','Vietnam':'APAC','Cambodia':'APAC','Laos':'APAC',
    'Indonesia':'APAC','Malaysia':'APAC','Singapore':'APAC','Philippines':'APAC',
    'China':'APAC','Korea':'APAC','South Korea':'APAC','Japan':'APAC',
    'Taiwan':'APAC','Hong Kong':'APAC',
    'Australia':'APAC','New Zealand':'APAC',
    # ─────── AMER · Apocalypse −3 to −10 (Americas) ───────
    'USA':'AMER','EEUU':'AMER','United States':'AMER',
    'Canada':'AMER','Mexico':'AMER','Guatemala':'AMER','Honduras':'AMER',
    'Costa Rica':'AMER','Panama':'AMER','Cuba':'AMER','Dominican Republic':'AMER',
    'Brazil':'AMER','Argentina':'AMER','Chile':'AMER','Peru':'AMER','Ecuador':'AMER',
    'Colombia':'AMER','Venezuela':'AMER','Uruguay':'AMER','Paraguay':'AMER','Bolivia':'AMER',
    # ─────── Unknown ───────
    'ONU':'UNK', '':'UNK', None:'UNK',
}
# Labels reference Apocalypse Time (UTC). The AD week closes at 00:00 Apocalypse.
ZONE_LABEL = {
    'EMEA': '🌍 EMEA · Apocalypse −1/+4',
    'APAC': '🌏 Asia-Pacific · Apocalypse +5/+12',
    'AMER': '🌎 Americas · Apocalypse −3/−10',
    'UNK':  '❓ Unknown',
}
DAYS_ES = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado']
DAYS_EN = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday']
ES_TO_EN = dict(zip(DAYS_ES, DAYS_EN))


def open_wb(path: Path):
    if not path.exists():
        raise FileNotFoundError(f"No encuentro {path}")
    return openpyxl.load_workbook(str(path), data_only=True, read_only=True)


# ───────────────────────────────────────────────────────────
# Normalización de país: distintas variantes ortográficas o
# idiomáticas se mapean a un nombre canónico único.
# ───────────────────────────────────────────────────────────
COUNTRY_ALIASES = {
    'Indien':  'India',        # variante alemana
    'EEUU':    'USA',          # estándar interno → ISO common
    'Türkiye': 'Turkey',
    None:      'ONU',
    '':        'ONU',
}
def normalize_country(c):
    return COUNTRY_ALIASES.get(c, c)


# ───────────────────────────────────────────────────────────
# 1) DETECTAR SEMANA EN CURSO (carpeta más reciente del ciclo)
# ───────────────────────────────────────────────────────────
def find_current_week_xlsx(cycle_dir_name: str = DEFAULT_CYCLE_DIR_NAME):
    """Devuelve (path, week_number, rival, cycle_label, week_folder_name)."""
    cycle_root = ROOT / cycle_dir_name
    if not cycle_root.exists():
        return None
    candidates = list(cycle_root.glob("*/Datos_parciales_*.xlsx"))
    if not candidates:
        return None
    most_recent = max(candidates, key=lambda p: p.stat().st_mtime)
    folder = most_recent.parent.name  # ej. "01-WYN vs OM3N"
    # week number = parte numérica al inicio de la carpeta
    wn = 1
    for tok in folder.replace('-', ' ').split():
        if tok.isdigit():
            wn = int(tok); break
    # rival = lo que viene tras "vs " en el nombre de la carpeta
    rival = folder.split('vs', 1)[1].strip() if 'vs' in folder else '?'
    return most_recent, wn, rival, cycle_dir_name, folder


# ───────────────────────────────────────────────────────────
# 2) LECTURA DE Datos_parciales (fuente PRIMARIA mid-week)
# ───────────────────────────────────────────────────────────
def extract_from_datos_parciales(path: Path):
    """
    Devuelve:
      players_by_day:  {nombre: {Lunes: {points,rank}, ...}}     ← solo WYN
      day_team_pts:    {Lunes: total_pts WYN, ...}
      day_team_active: {Lunes: n_active WYN, ...}
      day_rival_pts:   {Lunes: total_pts rival, ...}
      days_with_data:  ['Lunes','Martes',...] en orden
    """
    wb = open_wb(path)
    by_day = {}
    day_team_pts = {d: 0 for d in DAYS_ES}
    day_team_active = {d: 0 for d in DAYS_ES}
    day_rival_pts = {d: 0 for d in DAYS_ES}
    days_with_data = []
    for sh in DAYS_ES:
        if sh not in wb.sheetnames:
            continue
        rows = list(wb[sh].iter_rows(values_only=True))
        header_idx = None
        for i, r in enumerate(rows[:5]):
            if r and r[0] == 'Pos':
                header_idx = i; break
        if header_idx is None:
            continue
        has_data = False
        for r in rows[header_idx+1:]:
            if not r[0] or not r[1] or not r[3]:
                continue
            pos, name, pts, team = r[:4]
            pts = pts or 0
            team_s = str(team)
            # Skip subtotal rows: Excel adds rows like '⚔️ OM3N +25,920,630' which
            # are aggregates already included via the individual player rows.
            if '+' in team_s:
                continue
            if 'WYN' in team_s:
                by_day.setdefault(name, {})[sh] = {'rank': pos, 'points': pts}
                day_team_pts[sh] += pts
                if pts > 0:
                    day_team_active[sh] += 1
                    has_data = True
            else:
                # rival (cualquier otro equipo)
                day_rival_pts[sh] += pts
                if pts > 0:
                    has_data = True
        if has_data:
            days_with_data.append(sh)
    return by_day, day_team_pts, day_team_active, day_rival_pts, days_with_data


# ───────────────────────────────────────────────────────────
# 3) LECTURA DEL MASTER (agregado oficial al cierre semanal)
# ───────────────────────────────────────────────────────────
def find_week_sheet(wb, week_number: int):
    """Busca la hoja 'WeekN WYN' (N = week_number) si existe."""
    target = f"week{week_number}".lower()
    for sh in wb.sheetnames:
        s = sh.lower().replace(' ', '')
        if s.startswith(target) and 'wyn' in s:
            return sh
    return None


def extract_master_ranking(week_number: int):
    """Si la hoja WeekN WYN existe y tiene datos, devuelve [{rank,name,points,pctOfTotal,country}]."""
    if not MASTER.exists():
        return None, None
    wb = open_wb(MASTER)
    sheet = find_week_sheet(wb, week_number)
    if not sheet:
        return None, None
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, r in enumerate(rows[:5]):
        if r and r[0] == 'Pos':
            header_idx = i; break
    if header_idx is None:
        return None, None
    title = rows[0][0] if rows[0] else ''
    rival = None
    if isinstance(title, str) and 'vs' in title:
        rival = title.split('vs')[1].split('—')[0].strip()
    players = []
    for r in rows[header_idx+1:]:
        if r[0] == 'TOTAL' or not isinstance(r[0], int):
            continue
        pos, name, points, pct, country = (r + (None,)*5)[:5]
        players.append({
            'rank': pos, 'name': name, 'points': points or 0,
            'pctOfTotal': pct, 'country': normalize_country(country or 'ONU'),
        })
    return (players or None), rival


def extract_master_cofres():
    """Devuelve {nombre: {gained, possible, percentage}} de la hoja Cofres — Ranking."""
    if not MASTER.exists():
        return {}
    wb = open_wb(MASTER)
    if 'Cofres — Ranking' not in wb.sheetnames:
        return {}
    out = {}
    for r in wb['Cofres — Ranking'].iter_rows(values_only=True):
        if isinstance(r[0], int) and isinstance(r[2], str) and '/' in r[2]:
            g, p = [int(x.strip()) for x in r[2].split('/')]
            out[r[1]] = {'gained': g, 'possible': p, 'percentage': round((r[3] or 0)*100, 2)}
    return out


def extract_datos_cofres_blocks(xlsx_path: Path):
    """Devuelve los bloques ▌ Semana N de la hoja Datos Cofres con KPIs del equipo por día."""
    if not xlsx_path.exists():
        return []
    wb = open_wb(xlsx_path)
    if 'Datos Cofres' not in wb.sheetnames:
        return []
    rows = list(wb['Datos Cofres'].iter_rows(values_only=True))
    out = []
    for i, r in enumerate(rows):
        c = r[0]
        if isinstance(c, str) and c.startswith('▌'):
            j = i + 1
            while j < len(rows) and rows[j][0] != 'Métrica':
                j += 1
            if j < len(rows):
                metrics = rows[j+1:j+6]
                if len(metrics) < 5: continue
                activos, puntos, verde, media_p, diam = metrics
                by_day = {}
                for k, d in enumerate(DAYS_ES, start=1):
                    by_day[d] = {
                        'active': activos[k], 'points': puntos[k],
                        'pctGreen': verde[k], 'avgCoffersLost': media_p[k],
                    }
                out.append({
                    'label': c,
                    'byDay': by_day,
                    'totalPoints': puntos[7] or sum(x['points'] or 0 for x in by_day.values()),
                    'avgPctGreenWeek': verde[7],
                    'avgCoffersLostPerPlayer': media_p[7],
                    'diamondsLost': diam[1],
                    'diamondsPossible': diam[3],
                    'diamondsLostPct': (diam[1]/diam[3]) if (diam[1] and diam[3]) else None,
                })
    return out


# ───────────────────────────────────────────────────────────
# 4) PIPELINE PRINCIPAL
# ───────────────────────────────────────────────────────────
def build_current_week():
    """Construye el JSON de la semana actual. Devuelve (data, current_block_kpi, dataSource, warnings)."""
    warnings = []

    info = find_current_week_xlsx()
    if not info:
        raise RuntimeError(f"No encuentro carpeta '{DEFAULT_CYCLE_DIR_NAME}/0X-WYN vs RIVAL/Datos_parciales_*.xlsx'")
    parciales_path, week_number, rival_raw, cycle_label, folder_name = info

    # 4.1) Leer Datos_parciales (vivo)
    by_day, day_team_pts, day_team_active, day_rival_pts, days_with_data = extract_from_datos_parciales(parciales_path)

    # 4.2) Master: Week N (puede no existir aún si la semana no se ha cerrado)
    master_players, master_rival = extract_master_ranking(week_number)
    use_master = master_players is not None and len(master_players) > 0

    # 4.3) KPIs del equipo (Datos Cofres del master). Si la semana no se ha cerrado, no existirá.
    master_blocks = extract_datos_cofres_blocks(MASTER)
    current_block = None
    for blk in master_blocks:
        # match por número de semana o por rival
        if f'Semana {week_number}' in blk['label'] or (rival_raw and rival_raw.split(']')[-1][:4].lower() in blk['label'].lower()):
            current_block = blk; break
    if current_block is None and master_blocks:
        current_block = master_blocks[-1]  # fallback: el último bloque del master

    cofres = extract_master_cofres()

    # 4.4) Construir lista de jugadores
    players_raw = []
    if use_master:
        # Master tiene Week N — usar como ranking oficial
        for p in master_players:
            players_raw.append(p)
        data_source = 'master'
    else:
        # Live: agregar Lun-hoy desde Datos_parciales
        # Necesitamos también el país. Lo tomamos del último cierre conocido (master Week N-1),
        # o del Historial si existe.
        # Por simplicidad: si el master tiene Week (N-1), usar país de ahí.
        country_by_name = {}
        if MASTER.exists():
            wb_m = open_wb(MASTER)
            prev_sheet = find_week_sheet(wb_m, week_number - 1) if week_number > 1 else find_week_sheet(wb_m, week_number)
            if prev_sheet:
                rows = list(wb_m[prev_sheet].iter_rows(values_only=True))
                for r in rows:
                    if isinstance(r[0], int) and r[1]:
                        country_by_name[r[1]] = normalize_country(r[4] if len(r) > 4 and r[4] else 'ONU')
        # Construir totales
        totals = []
        for name, day_data in by_day.items():
            total = sum(d['points'] for d in day_data.values())
            totals.append({
                'name': name,
                'points': total,
                'country': country_by_name.get(name, 'ONU'),
            })
        totals.sort(key=lambda x: -x['points'])
        for i, t in enumerate(totals, start=1):
            t['rank'] = i
            t['pctOfTotal'] = None
        players_raw = totals
        data_source = 'live'
        warnings.append(f"Semana {week_number} aún no cerrada en el master. Mostrando agregado vivo Lun-{days_with_data[-1] if days_with_data else '?'} desde Datos_parciales.")

    if not players_raw:
        raise RuntimeError("No tengo jugadores. ¿El Excel está vacío?")

    total_pts = sum(p['points'] for p in players_raw)
    avg_pts = round(total_pts / len(players_raw)) if players_raw else 0

    # media diaria del equipo a partir de los datos vivos
    team_day_avg = {d: (day_team_pts[d]/day_team_active[d] if day_team_active[d] else 0) for d in DAYS_ES}

    # 4.5) Construir cada jugador con su desglose
    players = []
    unknown_countries = set()
    zone_counts = {}
    for p in players_raw:
        z = COUNTRY_ZONE.get(p['country'])
        if z is None:
            unknown_countries.add(p['country']); z = 'UNK'
        zone_counts[z] = zone_counts.get(z, 0) + 1
        cof = cofres.get(p['name'], {'gained': 0, 'possible': 54, 'percentage': 0})
        dd = by_day.get(p['name'], {})
        breakdown, daily_pts = [], []
        for es, en in zip(DAYS_ES, DAYS_EN):
            info = dd.get(es, {'rank': None, 'points': 0})
            avg = team_day_avg[es]
            delta = (info['points']/avg - 1)*100 if (avg and info['points']) else None
            breakdown.append({
                'day': en, 'dayName': en, 'dayNumber': DAYS_ES.index(es),
                'points': info['points'], 'rank': info['rank'],
                'percentVsAvg': round(delta, 1) if delta is not None else None,
                'cofferGained': None,
            })
            if info['points']: daily_pts.append((es, info['points']))
        if daily_pts:
            best  = max(daily_pts, key=lambda x: x[1])
            worst = min(daily_pts, key=lambda x: x[1])
            avg_d = sum(x[1] for x in daily_pts)/len(daily_pts)
            vol = (statistics.pstdev([x[1] for x in daily_pts])/avg_d*100) if avg_d else 0
            day_analysis = {
                'bestDay':  {'day': ES_TO_EN[best[0]],  'dayNumber': DAYS_ES.index(best[0]),  'points': best[1]},
                'worstDay': {'day': ES_TO_EN[worst[0]], 'dayNumber': DAYS_ES.index(worst[0]), 'points': worst[1]},
                'averageDaily': round(avg_d), 'volatility': round(vol, 1),
            }
        else:
            day_analysis = None
        vs_pct = round((p['points']/avg_pts - 1)*100, 1) if avg_pts else 0
        insights = []
        if cof['percentage'] >= 100: insights.append({'type':'positive','icon':'✅','message':'Perfect coffers (100%)'})
        elif cof['percentage'] and cof['percentage'] < 50: insights.append({'type':'warning','icon':'⚠️','message':f"Low coffers ({cof['percentage']:.0f}%)"})
        if vs_pct >= 100: insights.append({'type':'positive','icon':'⭐','message':f'Outstanding performance (+{vs_pct:.0f}%)'})
        elif vs_pct <= -50: insights.append({'type':'warning','icon':'⚠️','message':f'Low performance ({vs_pct:.0f}%)'})
        if day_analysis: insights.append({'type':'pattern','icon':'📊','message':f"Best day: {day_analysis['bestDay']['day']}"})
        players.append({
            'rank': p['rank'], 'playerId': p['name'], 'gameName': p['name'],
            'country': p['country'], 'timezone': {'zone': z, 'label': ZONE_LABEL[z]},
            'points': p['points'], 'vsAverage': {'points': p['points']-avg_pts, 'percentage': vs_pct},
            'coffers': {**cof, 'rank': None},
            'dayAnalysis': day_analysis, 'dailyBreakdown': breakdown, 'insights': insights,
        })
    for i, p in enumerate(sorted(players, key=lambda x: -x['coffers']['percentage']), start=1):
        p['coffers']['rank'] = i

    # 4.6) Estimar cofres por día (solo si tenemos current_block con KPIs por día)
    if current_block:
        for d_es, d_en in zip(DAYS_ES, DAYS_EN):
            info = current_block['byDay'][d_es]
            pct_g = info['pctGreen'] or 0
            avg_l = info['avgCoffersLost'] or 0
            actives, inactives = [], []
            for pl in players:
                pts = next(b['points'] for b in pl['dailyBreakdown'] if b['day']==d_en)
                (actives if pts > 0 else inactives).append(pl)
            actives.sort(key=lambda pl: -next(b['points'] for b in pl['dailyBreakdown'] if b['day']==d_en))
            n_perf = round(pct_g * len(actives))
            rest = actives[n_perf:]
            total_lost = avg_l * len(actives)
            if rest:
                N = len(rest)
                g_list = [round(8*(1 - i/(N-1))) if N > 1 else 4 for i in range(N)]
                cur_lost = sum(9-g for g in g_list)
                if cur_lost > 0 and total_lost > 0:
                    sc = total_lost/cur_lost
                    g_list = [max(0, min(9, round(9-(9-g)*sc))) for g in g_list]
            else:
                g_list = []
            for pl in actives[:n_perf]:
                for b in pl['dailyBreakdown']:
                    if b['day']==d_en: b['cofferGained']=9; break
            for pl, g in zip(rest, g_list):
                for b in pl['dailyBreakdown']:
                    if b['day']==d_en: b['cofferGained']=g; break
            for pl in inactives:
                for b in pl['dailyBreakdown']:
                    if b['day']==d_en: b['cofferGained']=None; break

    rival_clean = master_rival or rival_raw

    # Build per-day team totals + derive daily AD result (W/L by comparing WYN vs rival points)
    ES_TO_EN_MAP = dict(zip(DAYS_ES, DAYS_EN))
    wyn_daily   = {ES_TO_EN_MAP[d]: day_team_pts[d]   for d in DAYS_ES}
    rival_daily = {ES_TO_EN_MAP[d]: day_rival_pts[d]  for d in DAYS_ES}
    daily_results = {}
    for d in DAYS_ES:
        en = ES_TO_EN_MAP[d]
        w = day_team_pts[d]; r = day_rival_pts[d]
        if w == 0 and r == 0:
            daily_results[en] = None  # día sin datos
        elif w > r:
            daily_results[en] = 'W'
        elif w < r:
            daily_results[en] = 'L'
        else:
            daily_results[en] = 'D'  # empate (raro)

    data = {
        'weekId': f'GG_Mayo_W{week_number}',
        'cycleId': 'Gold_Group_Mayo',
        'weekNumber': week_number,
        'rival': rival_clean,
        'competitionDates': {'startDate': None, 'endDate': None},
        'sourceFolder': folder_name,
        'metadata': {
            'totalPlayers': len(players),
            'totalPoints': total_pts,
            'avgPoints': avg_pts,
            'playersByZone': zone_counts,
            'cofferGainedIsEstimate': True,
            'dataSource': data_source,
            'daysWithData': [ES_TO_EN[d] for d in days_with_data],
            'generatedAt': datetime.datetime.now().isoformat(timespec='seconds'),
        },
        'players': players,
        'wynDailyPoints':   wyn_daily,
        'rivalDailyPoints': rival_daily,
        'dailyResults':     daily_results,
    }
    return data, current_block, data_source, warnings, unknown_countries


def build_history(current_block):
    abril = extract_datos_cofres_blocks(ABRIL)
    mayo  = [current_block] if current_block else []

    baseline = None
    if abril:
        baseline = {
            'sourceWeeks': [f'Abril S{i+1}' for i in range(len(abril))],
            'n': len(abril),
            'avgTotalPoints': statistics.mean(w['totalPoints'] for w in abril),
            'avgPctGreenWeek': statistics.mean(w['avgPctGreenWeek'] for w in abril),
            'avgDiamondsLostPct': statistics.mean(w['diamondsLostPct'] for w in abril),
            'avgPointsPerDay': {d: statistics.mean(w['byDay'][d]['points'] for w in abril) for d in DAYS_ES},
            'avgPctGreenPerDay': {d: statistics.mean(w['byDay'][d]['pctGreen'] for w in abril) for d in DAYS_ES},
            'avgCoffersLostPerDay': {d: statistics.mean(w['byDay'][d]['avgCoffersLost'] for w in abril) for d in DAYS_ES},
        }

    return {
        'cycles': [
            {
                'cycleId': 'Diamond_Group_Abril', 'label': 'Diamond Group Abril',
                'division': 'Diamond', 'status': 'closed',
                'narrative': ('Subida a Diamond. Perdimos las cuatro semanas del ciclo y descendimos a Gold. '
                              f"Total acumulado: {sum(w['totalPoints'] for w in abril)/1e9:.2f} B pts." if abril else 'Sin datos'),
                'weeks': [
                    {'weekNumber': i+1,
                     'rival': w['label'].split('vs ')[1].split('(')[0].strip(),
                     'result': 'L',
                     **{k:v for k,v in w.items() if k!='label'}}
                    for i, w in enumerate(abril)
                ],
            },
            {
                'cycleId': 'Gold_Group_Mayo', 'label': 'Gold Group Mayo',
                'division': 'Gold', 'status': 'in_progress',
                'narrative': 'Vuelta a Gold tras Diamond Abril sin victorias. S1 vs OM3N perdida.',
                'weeks': [
                    {'weekNumber': i+1,
                     'rival': w['label'].split('vs ')[1].split('(')[0].strip(),
                     'result': 'L' if i == 0 else None,
                     **{k:v for k,v in w.items() if k!='label'}}
                    for i, w in enumerate(mayo)
                ],
            },
        ],
        'baseline': baseline,
    }


def _translate_day_keys(obj):
    """Recursively translate Spanish day keys to English in any nested dict."""
    DAY_TR = {'Lunes':'Monday','Martes':'Tuesday','Miércoles':'Wednesday',
              'Jueves':'Thursday','Viernes':'Friday','Sábado':'Saturday'}
    if isinstance(obj, dict):
        return {DAY_TR.get(k,k): _translate_day_keys(v) for k,v in obj.items()}
    if isinstance(obj, list):
        return [_translate_day_keys(x) for x in obj]
    return obj


def regenerate_html(data, history):
    """Refresh JSON-embedded HTML files.
       index.html      → rebuilt from template.html if available
       dashboard.html  → in-place re-embedding of WYN_DATA / WYN_HISTORY
       v2-preview.html → in-place re-embedding of WYN_DATA / WYN_HISTORY"""
    def embed(name, obj):
        return f'window.{name} = ' + json.dumps(obj, ensure_ascii=True).replace('</', '<\\/') + ';'

    any_written = False

    # 1) index.html: rebuilt from template
    if TEMPLATE.exists():
        out = TEMPLATE.read_text(encoding='utf-8')
        out = out.replace('/* WYN_DATA_PLACEHOLDER */',    embed('WYN_DATA',    data))
        out = out.replace('/* WYN_HISTORY_PLACEHOLDER */', embed('WYN_HISTORY', history))
        OUT_HTML.write_text(out, encoding='utf-8')
        print(f'  ✔ Rebuilt {OUT_HTML.name} from template')
        any_written = True
    else:
        print(f'  ⚠ {TEMPLATE.name} missing — skipping {OUT_HTML.name}')

    # 2) dashboard.html and v2-preview.html: in-place re-embed
    for fname in ('dashboard.html', 'v2-preview.html'):
        fpath = HERE / fname
        if not fpath.exists():
            continue
        s = fpath.read_text(encoding='utf-8')
        for varname, obj in [('WYN_DATA', data), ('WYN_HISTORY', history)]:
            START = f'window.{varname} = '
            i = s.find(START)
            if i < 0:
                continue
            j = s.find(';\n', i)
            if j < 0:
                continue
            s = s[:i] + embed(varname, obj) + s[j+1:]
        fpath.write_text(s, encoding='utf-8')
        print(f'  ✔ Re-embedded data into {fname}')
        any_written = True

    return any_written


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--preview', action='store_true')
    args = ap.parse_args()

    print('—'*60)
    print('WYN AD Tracker · refresco desde Excel')
    print('—'*60)
    print(f'Master:   {MASTER.name}  ({"existe" if MASTER.exists() else "no existe"})')

    info = find_current_week_xlsx()
    if info:
        path, wn, rival, cycle, folder = info
        print(f'Semana actual detectada: {cycle} · S{wn} · carpeta "{folder}"')
        print(f'  Datos_parciales:  {path.relative_to(ROOT)}')
    else:
        print('!! No detecto carpeta de semana en curso')

    print(f'Histórico Abril: {ABRIL.name}  ({"existe" if ABRIL.exists() else "no existe"})')

    data, current_block, data_source, warnings, unknown = build_current_week()
    history = build_history(current_block)

    print()
    print(f"Semana: {data['cycleId']} · S{data['weekNumber']} vs {data['rival']}")
    print(f"Fuente: {data_source}  (master = semana cerrada · live = mid-week desde Datos_parciales)")
    print(f"Días con datos: {data['metadata']['daysWithData']}")
    print(f"  Jugadores:      {len(data['players'])}")
    print(f"  Puntos totales: {data['metadata']['totalPoints']/1e9:.3f} B")
    print(f"  Top 5:")
    for p in data['players'][:5]:
        c = p['coffers']
        cof_str = f"  ·  cofres {c['gained']}/{c['possible']}" if c['possible'] else ""
        print(f"    #{p['rank']:>3}  {p['gameName']:<22}  {p['country']:<18}  "
              f"{p['points']/1e6:>6.1f} M{cof_str}")
    print(f"  Zonas: {data['metadata']['playersByZone']}")

    for w in warnings:
        print(f"  ⚠ {w}")
    if unknown:
        print(f"  ⚠ Países sin mapear → Desconocida: {sorted(unknown)}")

    if args.preview:
        print('\n--preview · no se ha escrito ningún archivo.')
        return

    # Ensure all internal day-keys are in English (Spanish keys would break the dashboards)
    data    = _translate_day_keys(data)
    history = _translate_day_keys(history)

    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    OUT_HIST.write_text(json.dumps(history, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
    print(f"\n✔ Escrito: {OUT_JSON.name}")
    print(f"✔ Escrito: {OUT_HIST.name}")

    if regenerate_html(data, history):
        print(f"✔ Regenerado: {OUT_HTML.name}")
        print(f"\nAbre la app: file://{OUT_HTML}")


if __name__ == '__main__':
    main()
